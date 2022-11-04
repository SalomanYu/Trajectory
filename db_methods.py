import os
import sqlite3
import xlrd
import openpyxl

from typing import NamedTuple
from dataclasses import dataclass
from datetime import date
from settings.config import ResumeGroup
import settings.database as database
from settings.tools import get_default_names, group_steps_to_resume


CURRENT_DATABASE_NAME = "TestResult.db"

class SimilarWay(NamedTuple):
    """Используется в методе нахождения самого популярного пути"""
    resumeId: str
    similarId: int

class DistinctWay(NamedTuple):
    similarId: int
    area: str
    resumeId: str
    title: str
    steps: list[str]

class Connection(NamedTuple):
    """Класс подключения к Базе данных"""
    db: sqlite3.Connection
    cursor: sqlite3.Cursor

@dataclass(slots=True)
class ProfessionStep:
    """Объект, представляющий собой отдельную строку БД, т.е этап в карьере"""
    title: str
    experiencePost: str
    experienceInterval: str
    experienceDuration: str
    branch: str
    subbranch: str
    weightInGroup: int
    level: int
    levelInGroup: int
    groupId: int
    area: str
    city: str
    generalExcepience: str
    specialization: str
    salary: str
    educationUniversity: str
    educationDirection: str
    educationYear: str # Потому что может быть перечисление
    languages: str
    skills: str
    advancedTrainingTitle: str
    advancedTrainingDirection: str
    advancedTrainingYear: str
    dateUpdate: date
    resumeId: str
    similarPathId: int | None = None

class Resume(NamedTuple):
    id: str # ссылка на резюме 
    steps: list[ProfessionStep]

def connect_to_db(db_name: str = CURRENT_DATABASE_NAME) -> Connection:
    db = sqlite3.connect(db_name)
    cursor = db.cursor()
    return Connection(db, cursor)


def get_most_popular_work_way(table_name: str, db_name: str = CURRENT_DATABASE_NAME)-> tuple[ProfessionStep]:
    """Обращаемся к базе, данные которой прошли через все этапы построения путей и ищем профессию
    в заголовке резюме или в предыдущих должностях соискателя.
    Ищем наиболее повторяющиеся индефикаторы одинаковых путей.

    Возвращаем генератор списка, каждый элемент которого представляет собой строку из БД, 
    которая в свою очередь является отдельным этапом в карьере. 
    Возвращаются все резюме, связанные similarPathId
    Отделяются этапы разных резюме по resumeId, т.е ссылке на резюме 
    """

    db, cursor = connect_to_db(db_name)
    query = f"""SELECT DISTINCT resumeId, similarPathId FROM {table_name};"""
    cursor.execute(query)
    data = (SimilarWay(*item) for item in cursor.fetchall())
    
    distinct_ways = {} # Словарь вида Ключ-ID пути: Значение-Количество повторений
    for item in data:
        if item.similarId in distinct_ways: distinct_ways[item.similarId] += 1
        else: distinct_ways[item.similarId] = 1
    most_popular_similar_id = max(distinct_ways, key=distinct_ways.get) # Берем ключ максимального значения в словаре
    
    query = f"""SELECT * FROM {table_name} WHERE similarPathId={most_popular_similar_id}"""
    cursor.execute(query)
    result = (ProfessionStep(*row[1:]) for row in cursor.fetchall())
    
    db.close()
    return result  


def find_profession_in_work_ways(profession_name: str, table_name: str, db_name: str = CURRENT_DATABASE_NAME)\
     -> tuple[ProfessionStep] | None:
    """Обращаемся к базе, данные которой прошли через все этапы построения путей и ищем профессию
    в заголовке резюме или в предыдущих должностях соискателя.
    Если такая профессия есть в БД, то вернется генератор списка, содержащий все совпадения с профессией
    Замечание: В поиске не учитывается пунктуация и регистр
    """
        
    db, cursor = connect_to_db(db_name)
    query = f"""SELECT * FROM {table_name} 
        WHERE title='{profession_name}' OR experiencePost='{profession_name}'"""
    cursor.execute(query)
    data = (ProfessionStep(*item[1:]) for item in cursor.fetchall())
    db.close()
    if data: 
        return data
    return None


def find_most_popular_way(resumes: ResumeGroup, area: str, profession_name:str, count: int = 3) -> list[DistinctWay]:
    distinct_ways: dict[DistinctWay, int] = {} # Словарь типа Айди пути: количество повторений
    for resume in resumes:
        step = resume.ITEMS[0]
        if profession_name == step.title and step.area == area:
            current_way = DistinctWay(
                    similarId=step.similarPathId, 
                    area=step.area,
                    resumeId=resume.ID,
                    title=step.title,
                    steps=' | '.join([step.experiencePost for step in resume.ITEMS]))
            if current_way in distinct_ways:distinct_ways[current_way] += 1
            else:distinct_ways[current_way] = 1
    
    most_popular_way = sorted(distinct_ways, key=distinct_ways.get, reverse=True)
    return [key for key in most_popular_way[:count]]


def main():
    areas = {'Ветеринария', 'Охрана и безопасность', 'бухгалтерия и налоги','Инвестиции, ценные бумаги и управление финансами', }
    # areas =  {'бухгалтерия и налоги'}
    resumes = group_steps_to_resume(data=database.get_all_resumes(table_name='New'))
    default_names, _ = get_default_names(areas=areas)
    # Собрали все стандарные наименования для каждого уровня и профобласти
    book = openpyxl.Workbook()
    sheet = book.active
    row = 3
    sheet.cell(row=row, column=1, value='id')
    sheet.cell(row=row, column=2, value='№ группы')
    sheet.cell(row=row, column=3, value='Профобласть')
    sheet.cell(row=row, column=4, value='Целевая профессия')
    sheet.cell(row=row, column=5, value='Уровень')
    sheet.cell(row=row, column=6, value='Шаги')
    sheet.cell(row=row, column=7, value='Ссылка')

    for default in default_names:
        ways = find_most_popular_way(area=default.area, profession_name=default.name, count=3, resumes=resumes)
        for way in ways:
            row += 1
            sheet.cell(row=row, column=1, value=row-1)
            sheet.cell(row=row, column=2, value=default.profID)
            sheet.cell(row=row, column=3, value=default.area)
            sheet.cell(row=row, column=4, value=default.name)
            sheet.cell(row=row, column=5, value=default.level)
            sheet.cell(row=row, column=6, value=way.steps)
            sheet.cell(row=row, column=7, value=way.resumeId)
            print(default.name)
    book.save(filename='Data/MostPopular.xlsx')        
if __name__ == "__main__":
    # Пример работы с методами
    main()
    # Получить список всех популярных путей
    # most_popular_way = get_most_popular_work_way(table_name="finance")
    # for item in most_popular_way:
    #     print(item)        
    
    # # Найти профессию среди этапов или заголовков резюме
    # result = find_profession_in_work_ways(profession_name="Аналитик", table_name="finance", db_name="TestResult.db")
    # for item in result:
    #     print(f"{item.title}:{item.experiencePost}{item.resumeId}")