from typing import NamedTuple
import openpyxl

from settings.config import *
import  settings.database as database


class PopularWay(NamedTuple):
    way_id: int
    count_resumes: int
    area: str # Профобласть


def find_most_popular_workWays(tablename: str, area:str, level: int = 0) -> list[PopularWay]:
    data = database.get_all_resumes(tablename)
    dict_of_ways_with_indential_resume : dict[int,set[str]] = {} 
    result : list[PopularWay] = []

    for step in data:
        if step.level == level and step.area == area:
            if step.similarPathId not in dict_of_ways_with_indential_resume:
                dict_of_ways_with_indential_resume[step.similarPathId] = {step.resumeId} 
            else:
                dict_of_ways_with_indential_resume[step.similarPathId].add(step.resumeId) # Так как это множество, то здесь не будет повторяющихся ссылок на резюме. Таким образом мы посчитаем количество самых популярных путей
    
    sorted_dict = sorted(dict_of_ways_with_indential_resume.items(), key=lambda x: len(x[1]), reverse=True)
    for index in range(4):
        result.append(PopularWay(way_id=sorted_dict[index][0], count_resumes=len(sorted_dict[index][1]), area=area))    
    
    return result[1:] # Костыль

def show_way_by_similarPathId(tablename: str, similarPathId: int, area: str) -> tuple[ProfessionStep]:
    data = database.get_all_resumes(tablename)
    id_resume: str = [row.resumeId for row in data if row.similarPathId == similarPathId and row.area == area][0]
    data = database.get_all_resumes(tablename) # После одного использования генератор походу обнуляется, ибо в данный момент data пустой
    resume: tuple[ProfessionStep] = [row for row in data if row.resumeId == id_resume]
    return resume


def add_most_popular_professions_to_excel():
    excel_file = "Data/Популярные пути.xlsx"
    row = 1
    profareas = {'бухгалтерия и налоги', 'Инвестиции, ценные бумаги и управление финансами', 'Ветеринария', 'Охрана и безопасность'}
    popular_group_ways: tuple[list[PopularWay]] = (find_most_popular_workWays(tablename='New', level=level+1, area=area) for area in profareas for level in range(4))
    # popular_group_ways = [find_most_popular_workWays(tablename='New', level=level, area='Инвестиции, ценные бумаги и управление финансами')]

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row=row, column=1, value='id')    
    sheet.cell(row=row, column=2, value='Профобласть')
    sheet.cell(row=row, column=3, value='Количество')
    sheet.cell(row=row, column=4, value='Уровень')
    sheet.cell(row=row, column=5, value='Общий стаж')
    sheet.cell(row=row, column=6, value='Целевая профессия')
    sheet.cell(row=row, column=2, value='Должность')
    sheet.cell(row=row, column=2, value='Продолжительность')

    
    for group in popular_group_ways:
        for popular in group:
            resume = show_way_by_similarPathId(tablename='New', similarPathId=popular.way_id, area=popular.area)
            if not resume: break
            resume_basic = [step for step in resume][0] # Берем один элемент из генератора
            row += 1
            sheet.cell(row=row, column=1, value=row-1) # номер айди
            sheet.cell(row=row, column=2, value=resume_basic.area)
            sheet.cell(row=row, column=3, value=popular.count_resumes)
            sheet.cell(row=row, column=4, value=resume_basic.level)
            sheet.cell(row=row, column=5, value=resume_basic.generalExcepience)
            sheet.cell(row=row, column=6, value=resume_basic.title)
            for index, step in enumerate(resume):
                post = 'Нет опыта' if not step.experiencePost else step.experiencePost
                # duration = '' if not step.experienceDuration else step.experienceDuration
                branch = '-' if not step.branch else step.branch
                post_header = sheet.cell(row=1, column=7+index).value
                branch_header = sheet.cell(row=1, column=8+index).value
                # duration_header = sheet.cell(row=1, column=8+index).value
                if not post_header: sheet.cell(row=1, column=7+index, value=f"Должность {index+1}")
                if not branch_header: sheet.cell(row=1, column=8+index, value=f"Отрасль {index+1}")
                # if not duration_header: sheet.cell(row=1, column=8+index, value=f"Продолжительность {index+1}")
                sheet.cell(row=row, column=7+index, value=post)
                sheet.cell(row=row, column=8+index, value=branch)
                # sheet.cell(row=row, column=8+index, value=duration)
            print(resume_basic.title, resume_basic.level)

    book.save(excel_file)


def sort_steps_by_experience_interval(resumes: list[ResumeGroup]):
    ...

